"""One-shot import of the HubSpot Pipeline Book xlsx export.

Mirrors the logic of POST /api/prospects/import-xlsx. Safe to re-run:
creates prospects/relationships only if missing. Usage:

    python3 scripts/import_pipeline_book.py <path/to/Pipeline Book.xlsx>
"""
import os
import re
import sys
from urllib.parse import urlparse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.database import SessionLocal, engine, Base
from app.models import Prospect, Relationship


def domain_from_url(url):
    if not url:
        return None
    url = url.strip()
    if not url:
        return None
    if not re.match(r"^https?://", url, re.I):
        url = "http://" + url
    try:
        host = urlparse(url).hostname or ""
        return host.lower().lstrip("www.") or None
    except Exception:
        return None


def run(xlsx_path):
    Base.metadata.create_all(bind=engine)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    def cell(row, col_name):
        i = idx.get(col_name)
        if i is None or i >= len(row):
            return None
        v = row[i]
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    db = SessionLocal()
    prospects_created = 0
    relationships_created = 0
    rows_skipped = 0
    try:
        for row in rows:
            if not any(c is not None and str(c).strip() for c in row):
                continue
            company = cell(row, "Company")
            if not company:
                rows_skipped += 1
                continue

            prospect = db.query(Prospect).filter(Prospect.name.ilike(company)).first()
            if not prospect:
                prospect = Prospect(
                    name=company,
                    type=cell(row, "Partner Type"),
                    website=cell(row, "Company Website"),
                    domain=domain_from_url(cell(row, "Company Website")),
                )
                db.add(prospect)
                db.flush()
                prospects_created += 1

            email = cell(row, "Email")
            full_name = cell(row, "Full Name") or " ".join(
                filter(None, [cell(row, "First Name"), cell(row, "Last Name")])
            ) or None

            if email:
                existing = (
                    db.query(Relationship)
                    .filter(
                        Relationship.prospect_id == prospect.id,
                        Relationship.contact_email.ilike(email),
                    )
                    .first()
                )
                if existing:
                    rows_skipped += 1
                    continue
            elif full_name:
                existing = (
                    db.query(Relationship)
                    .filter(
                        Relationship.prospect_id == prospect.id,
                        Relationship.contact_name.ilike(full_name),
                    )
                    .first()
                )
                if existing:
                    rows_skipped += 1
                    continue
            else:
                rows_skipped += 1
                continue

            context_parts = []
            for label, col in [
                ("Pipeline Stage", "Pipeline Stage"),
                ("Lifecycle", "Lifecycle Stage"),
                ("Priority", "Priority"),
                ("My Contact", "My Contact"),
                ("BD Lead", "BD Team Lead"),
                ("HubSpot ID", "HubSpot ID"),
                ("Notes", "Source Notes"),
            ]:
                val = cell(row, col)
                if val:
                    context_parts.append(f"{label}: {val}")

            rel = Relationship(
                prospect_id=prospect.id,
                contact_name=full_name,
                contact_title=cell(row, "Title"),
                contact_email=email,
                contact_linkedin=cell(row, "LinkedIn URL"),
                score=2,
                context=" | ".join(context_parts) or None,
                source="pipeline_book",
            )
            db.add(rel)
            relationships_created += 1

        db.commit()
    finally:
        db.close()

    print(f"prospects_created:     {prospects_created}")
    print(f"relationships_created: {relationships_created}")
    print(f"rows_skipped:          {rows_skipped}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python3 scripts/import_pipeline_book.py <xlsx>")
        sys.exit(1)
    run(sys.argv[1])
