import csv
import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import ChecklistSet, ChecklistItem
from app.schemas import (
    ChecklistSetCreate, ChecklistSetOut, ChecklistSetUpdate, ChecklistItemIn, ChecklistItemOut,
)

router = APIRouter(prefix="/api/checklists", tags=["checklists"])


def _rows_from_xlsx(raw: bytes) -> list:
    """Read the first sheet of an .xlsx file into a list of string-cell rows."""
    import io as _io
    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            rows.append(cells)
    wb.close()
    return rows


def _set_summary(db: Session, s: ChecklistSet) -> dict:
    total = db.query(func.count(ChecklistItem.id)).filter(ChecklistItem.set_id == s.id).scalar() or 0
    owned = (
        db.query(func.count(ChecklistItem.id))
        .filter(ChecklistItem.set_id == s.id, ChecklistItem.owned.is_(True))
        .scalar()
        or 0
    )
    pct = round(owned / total * 100.0, 1) if total else 0.0
    return {
        "id": s.id, "name": s.name, "year": s.year, "brand": s.brand,
        "sport": s.sport, "notes": s.notes,
        "total_items": total, "owned_items": owned, "pct_complete": pct,
    }


@router.post("", response_model=ChecklistSetOut)
def create_set(payload: ChecklistSetCreate, db: Session = Depends(get_db)):
    data = payload.model_dump()
    if not (data.get("name") or "").strip():
        data["name"] = "Untitled checklist"
    s = ChecklistSet(**data)
    db.add(s)
    db.commit()
    db.refresh(s)
    return _set_summary(db, s)


@router.put("/{set_id}", response_model=ChecklistSetOut)
def update_set(set_id: int, payload: ChecklistSetUpdate, db: Session = Depends(get_db)):
    s = db.get(ChecklistSet, set_id)
    if not s:
        raise HTTPException(status_code=404, detail="Checklist not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(s, key, value)
    db.commit()
    db.refresh(s)
    return _set_summary(db, s)


@router.get("", response_model=list[ChecklistSetOut])
def list_sets(db: Session = Depends(get_db)):
    sets = db.query(ChecklistSet).order_by(ChecklistSet.id.desc()).all()
    return [_set_summary(db, s) for s in sets]


@router.get("/{set_id}")
def get_set(set_id: int, db: Session = Depends(get_db)):
    s = db.get(ChecklistSet, set_id)
    if not s:
        raise HTTPException(status_code=404, detail="Checklist not found")
    items = (
        db.query(ChecklistItem)
        .filter(ChecklistItem.set_id == set_id)
        .order_by(ChecklistItem.id)
        .all()
    )
    return {
        "set": _set_summary(db, s),
        "items": [ChecklistItemOut.model_validate(i).model_dump() for i in items],
    }


@router.delete("/{set_id}")
def delete_set(set_id: int, db: Session = Depends(get_db)):
    s = db.get(ChecklistSet, set_id)
    if not s:
        raise HTTPException(status_code=404, detail="Checklist not found")
    db.query(ChecklistItem).filter(ChecklistItem.set_id == set_id).delete()
    db.delete(s)
    db.commit()
    return {"ok": True}


@router.post("/{set_id}/items", response_model=ChecklistItemOut)
def add_item(set_id: int, payload: ChecklistItemIn, db: Session = Depends(get_db)):
    if not db.get(ChecklistSet, set_id):
        raise HTTPException(status_code=404, detail="Checklist not found")
    item = ChecklistItem(set_id=set_id, **payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.post("/{set_id}/import")
def import_items(set_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import checklist entries from a CSV. Recognized (case-insensitive) headers:
    card_number / number / #, player / name, variation / parallel, owned.
    A header row with none of those still works: col1 = card #, col2 = player.
    """
    if not db.get(ChecklistSet, set_id):
        raise HTTPException(status_code=404, detail="Checklist not found")

    raw = file.file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        try:
            rows = _rows_from_xlsx(raw)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {exc}")
    elif name.endswith(".xls"):
        raise HTTPException(
            status_code=400,
            detail="Old .xls files aren't supported — open it in Excel/Numbers and 'Save As' .xlsx or .csv.",
        )
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        rows = [r for r in csv.reader(io.StringIO(text)) if any(c.strip() for c in r)]
    if not rows:
        return {"imported": 0}

    def norm(h):
        return h.strip().lower()

    header = [norm(h) for h in rows[0]]
    num_aliases = {"card_number", "number", "card #", "card#", "#", "no", "no."}
    name_aliases = {"player", "name", "player name"}
    var_aliases = {"variation", "parallel", "insert", "version"}
    owned_aliases = {"owned", "have", "got"}

    has_header = any(h in num_aliases | name_aliases | var_aliases | owned_aliases for h in header)
    if has_header:
        idx = {
            "num": next((i for i, h in enumerate(header) if h in num_aliases), None),
            "name": next((i for i, h in enumerate(header) if h in name_aliases), None),
            "var": next((i for i, h in enumerate(header) if h in var_aliases), None),
            "owned": next((i for i, h in enumerate(header) if h in owned_aliases), None),
        }
        data_rows = rows[1:]
    else:
        idx = {"num": 0, "name": 1, "var": 2, "owned": None}
        data_rows = rows

    def cell(row, i):
        return row[i].strip() if i is not None and i < len(row) else ""

    truthy = {"1", "true", "yes", "y", "x", "owned", "have"}
    count = 0
    for row in data_rows:
        item = ChecklistItem(
            set_id=set_id,
            card_number=cell(row, idx["num"]),
            player=cell(row, idx["name"]),
            variation=cell(row, idx["var"]),
            owned=cell(row, idx["owned"]).lower() in truthy if idx["owned"] is not None else False,
        )
        db.add(item)
        count += 1
    db.commit()
    return {"imported": count}


@router.put("/items/{item_id}", response_model=ChecklistItemOut)
def update_item(item_id: int, payload: ChecklistItemIn, db: Session = Depends(get_db)):
    item = db.get(ChecklistItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/items/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db)):
    item = db.get(ChecklistItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()
    return {"ok": True}
