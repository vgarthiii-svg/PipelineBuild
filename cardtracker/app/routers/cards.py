import csv
import io
import os
import re
import uuid
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db, IMAGE_DIR
from app.models import Card, VALID_STATUSES, STATUS_IN_STOCK, STATUS_SOLD
from app.schemas import CardCreate, CardOut, CardUpdate, ExtractResult, UploadResult
from app.vision import extract_card

router = APIRouter(prefix="/api/cards", tags=["cards"])

_ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def _save_upload(upload: UploadFile) -> str:
    """Save an uploaded image under data/images and return its stored filename."""
    ext = os.path.splitext(upload.filename or "")[1].lower()
    if ext not in _ALLOWED_EXT:
        # Fall back to jpeg if the client didn't give a usable extension
        ext = ".jpg"
    name = f"{uuid.uuid4().hex}{ext}"
    dest = os.path.join(IMAGE_DIR, name)
    with open(dest, "wb") as f:
        f.write(upload.file.read())
    return name


def _max_card_num(db: Session) -> int:
    max_n = 0
    for (cid,) in db.query(Card.card_id).all():
        m = re.match(r"INV-(\d+)", cid or "", re.IGNORECASE)
        if m:
            max_n = max(max_n, int(m.group(1)))
    return max_n


def _next_card_id(db: Session) -> str:
    """Compute the next INV-#### id by scanning existing ids."""
    return f"INV-{_max_card_num(db) + 1:04d}"


# Column header (lowercased) -> Card field. Covers our own export plus common
# marketplace/eBay sales-log headers (e.g. "Card", "Sold Price", "eBay Fees").
_IMPORT_MAP = {
    # card identity / title
    "player": "player", "card": "player", "card name": "player", "title": "player",
    "description": "player", "card description": "player", "item": "player",
    "year": "year", "brand": "brand",
    "set": "set_name", "set name": "set_name", "set_name": "set_name",
    "card #": "card_number", "card number": "card_number", "number": "card_number",
    "card_number": "card_number", "#": "card_number",
    "variation": "variation", "parallel": "variation",
    "team": "team", "sport": "sport", "rookie": "is_rookie", "condition": "condition",
    "status": "status",
    # dates
    "date": "sale_date", "sale date": "sale_date", "purchase date": "purchase_date",
    # cost basis
    "purchase source": "purchase_source", "purchased from": "purchase_source",
    "purchase price": "purchase_price", "cost": "purchase_price",
    "inventory expense": "purchase_price", "cost basis": "purchase_price",
    # sale
    "sale platform": "sale_platform", "sold on": "sale_platform",
    "sale price": "sale_price", "sold price": "sale_price",
    "sale fees": "sale_fees", "fees": "sale_fees", "ebay fees": "sale_fees",
    "sale shipping": "sale_shipping",
    "est. value": "estimated_value", "estimated value": "estimated_value",
    "notes": "notes", "ebay item id": "notes", "item id": "notes", "listing id": "notes",
    "front image": "front_image", "back image": "back_image",
}
_FLOAT_FIELDS = {"purchase_price", "sale_price", "sale_fees", "sale_shipping", "estimated_value"}


def _parse_money(val: str):
    """Parse a money cell to a positive magnitude. Handles $, commas, and
    accounting negatives like ($9.33). Returns None for blanks/dashes."""
    s = re.sub(r"[,$()\s]", "", str(val)).replace("−", "")
    if s in ("", "-", "—", "–"):
        return None
    try:
        return abs(float(s))
    except ValueError:
        return None


# Best-effort extraction of a player name / year / brand from a freeform card
# title like "2023 Panini Prizm Football PUKA NACUA Orange ... RC LA Rams".
_TITLE_NOISE = set("""
topps panini bowman leaf donruss fleer upper deck upperdeck score sage onyx
gallery sapphire chrome prizm prizms select mosaic optic contenders absolute phoenix
obsidian spectra immaculate national treasures heritage stadium club finest fire
metal sports heroes garbage pail kids gpk series update concourse premier xrc
prospects adios reactive
refractor xfractor wave lime yellow green orange silver black red blue gold pink purple
teal bronze snowflake winter lazer laser shock zebra cracked ice holo mirror
golden emergent selections scripts huddle color match parallel insert rookie
ticket patch auto autograph rc sp ssp sealed rare lot disco velocity hyper fast
break mania downtown kaboom football basketball baseball hockey
psa bgs sgc gem mint hof rpm cgc nfl mlb nba nhl mls team card of and the no with x na a b
""".split())
_TITLE_SUFFIX = {"jr", "sr", "ii", "iii", "iv", "v"}
_TITLE_TEAMS = set("""
braves tigers dodgers spurs browns packers rams jaguars bears cardinals angels
mariners cavaliers grizzlies pelicans knicks lakers celtics warriors nuggets heat
bucks suns kings clippers mavericks rockets jazz thunder timberwolves pistons
pacers hawks hornets magic wizards raptors sixers blazers trailblazers
yankees redsox mets phillies cubs whitesox reds brewers pirates nationals
marlins astros rangers athletics twins royals guardians indians orioles bluejays
rays diamondbacks dbacks padres giants rockies
chiefs raiders broncos chargers bills dolphins patriots jets steelers ravens bengals
titans colts texans vikings lions eagles cowboys commanders niners seahawks saints
falcons panthers buccaneers bucs oilers flames canucks kraken sharks ducks avalanche
detroit chicago boston green bay los angeles la new york san antonio francisco diego
kansas city tampa
""".split())
_KNOWN_BRANDS = ["Topps", "Panini", "Bowman", "Leaf", "Donruss", "Fleer", "Upper Deck", "Score"]


def _format_player(items: list) -> str:
    out, seen = [], set()
    for tc, lt in items:
        if lt in seen:  # stop on a repeat, e.g. "Cam Schlittler, Cam Schlittler"
            break
        seen.add(lt)
        if lt in _TITLE_SUFFIX:
            out.append(lt.upper() if lt in {"ii", "iii", "iv", "v"} else lt.capitalize() + ".")
        elif tc.isupper():
            out.append(tc.capitalize())  # JAYDEN -> Jayden
        else:
            out.append(tc)               # preserve LaDainian / Jayden
    return " ".join(out).strip()


def _parse_card_title(desc: str) -> dict:
    """Pull {player, year, brand} out of a freeform card description (best effort)."""
    ym = re.search(r"\b(19|20)\d{2}\b", desc)
    year = ym.group(0) if ym else ""
    low = desc.lower()
    brand = next((b for b in _KNOWN_BRANDS if b.lower() in low), "")
    toks = []
    for t in desc.replace(",", " ").split():
        tc = t.strip().strip(".").strip("-")
        if not tc or any(ch.isdigit() for ch in tc) or tc.startswith("#") or "/" in tc:
            continue
        lt = re.sub(r"[^a-z]", "", tc.lower())
        if lt:
            toks.append((tc, lt))
    runs, cur = [], []
    for tc, lt in toks:
        if lt not in _TITLE_NOISE and lt not in _TITLE_TEAMS:
            cur.append((tc, lt))
        elif lt in _TITLE_SUFFIX and cur:
            cur.append((tc, lt))
        else:
            if cur:
                runs.append(cur)
                cur = []
    if cur:
        runs.append(cur)
    player = ""
    for run in runs:
        if len([1 for _, lt in run if lt not in _TITLE_SUFFIX]) >= 2:
            player = _format_player(run[:4])
            break
    return {"player": player, "year": year, "brand": brand}


def _enrich_from_title(data: dict):
    """When a row's 'player' is really a full card title, split out player/year/brand."""
    title = (data.get("player") or "").strip()
    if not title or data.get("year"):
        return
    if len(title.split()) < 4 and not re.search(r"\b(19|20)\d{2}\b", title):
        return  # short, real Player value (e.g. our own template) — leave it alone
    parsed = _parse_card_title(title)
    if parsed["player"]:
        data["player"] = parsed["player"]
        note = (data.get("notes") or "").strip()
        data["notes"] = (note + " — " if note else "") + title  # keep the full original
    if parsed["year"] and not data.get("year"):
        data["year"] = parsed["year"]
    if parsed["brand"] and not data.get("brand"):
        data["brand"] = parsed["brand"]


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime) or isinstance(v, date):
        return v.strftime("%m/%d/%Y")
    return str(v)


def _read_import_rows(upload: UploadFile) -> list:
    """Return non-empty rows (lists of string cells) from a CSV or Excel (.xlsx) upload."""
    raw = upload.file.read()
    name = (upload.filename or "").lower()
    if name.endswith(".xlsx") or raw[:4] == b"PK\x03\x04":  # xlsx is a zip (PK header)
        from openpyxl import load_workbook
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(
                status_code=400,
                detail="Couldn't read that spreadsheet. If it's an Apple Numbers file, "
                       "open it and choose File → Export To → Excel, then import the .xlsx.",
            )
        rows = []
        for r in wb.active.iter_rows(values_only=True):
            cells = [_cell_str(c) for c in r]
            if any(c.strip() for c in cells):
                rows.append(cells)
        wb.close()
        return rows
    text = raw.decode("utf-8-sig", errors="replace")
    return [r for r in csv.reader(io.StringIO(text)) if any(c.strip() for c in r)]


@router.post("/import")
def import_cards(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import inventory from a CSV or Excel (.xlsx) file. Auto-detects the header
    row, skipping any title/blank rows above it (common in eBay/marketplace exports)."""
    rows = _read_import_rows(file)
    if not rows:
        return {"imported": 0}
    # Find the header row: the first of the top rows that has >=2 known columns.
    header_idx, cols = 0, {}
    for idx in range(min(20, len(rows))):
        hdr = [str(h).strip().lower() for h in rows[idx]]
        candidate = {i: _IMPORT_MAP[h] for i, h in enumerate(hdr) if h in _IMPORT_MAP}
        if len(candidate) >= 2:
            header_idx, cols = idx, candidate
            break
    if not cols:
        raise HTTPException(
            status_code=400,
            detail="No recognizable columns found. Make sure the sheet has a header row "
                   "with names like Card/Player, Sold Price, Inventory Expense, etc.",
        )

    n = _max_card_num(db)
    count = 0
    for row in rows[header_idx + 1:]:
        data = {}
        for i, field in cols.items():
            if i >= len(row):
                continue
            val = row[i].strip()
            if not val:
                continue
            if field in _FLOAT_FIELDS:
                money = _parse_money(val)
                if money is not None:
                    data[field] = money
            else:
                data[field] = val
        if not any(data.get(k) for k in ("player", "year", "brand", "set_name", "card_number")):
            continue
        _enrich_from_title(data)
        status = (data.pop("status", "") or "").lower().replace(" ", "_")
        if status not in VALID_STATUSES:
            # No status column: a row with a sale price is a sold card, else in stock.
            status = STATUS_SOLD if data.get("sale_price") is not None else STATUS_IN_STOCK
        n += 1
        db.add(Card(card_id=f"INV-{n:04d}", status=status, **data))
        count += 1
    db.commit()
    return {"imported": count}


@router.post("/upload", response_model=UploadResult)
def upload(
    front: UploadFile = File(...),
    back: UploadFile = File(None),
):
    """
    Free path (no AI, no cost): save the front/back photos and return their
    filenames. The browser handles any on-device text scanning itself; this
    endpoint just stores the images.
    """
    front_name = _save_upload(front)
    back_name = _save_upload(back) if back is not None else ""
    return UploadResult(front_image=front_name, back_image=back_name)


@router.post("/extract", response_model=ExtractResult)
def extract(
    front: UploadFile = File(...),
    back: UploadFile = File(None),
):
    """
    Step 1 of adding a card: upload front (and optional back) photos.
    Saves the images and runs Claude Vision to pre-fill the card's fields.
    The images are kept; pass their filenames back to POST /api/cards to save.
    """
    front_name = _save_upload(front)
    back_name = _save_upload(back) if back is not None else ""

    result = extract_card(
        os.path.join(IMAGE_DIR, front_name),
        os.path.join(IMAGE_DIR, back_name) if back_name else None,
    )
    return ExtractResult(
        front_image=front_name,
        back_image=back_name,
        extracted=result["fields"],
        vision_used=result["vision_used"],
        message=result["message"],
    )


@router.post("", response_model=CardOut)
def create_card(payload: CardCreate, db: Session = Depends(get_db)):
    """Step 2: save a reviewed card to inventory. Auto-assigns the next INV id."""
    status = payload.status if payload.status in VALID_STATUSES else "in_stock"
    card = Card(
        card_id=_next_card_id(db),
        status=status,
        front_image=payload.front_image,
        back_image=payload.back_image,
        purchase_price=payload.purchase_price,
        estimated_value=payload.estimated_value,
        purchase_date=payload.purchase_date,
        purchase_source=payload.purchase_source,
        **payload.model_dump(
            include={
                "player", "year", "brand", "set_name", "card_number",
                "variation", "team", "sport", "is_rookie", "condition", "notes",
            }
        ),
    )
    db.add(card)
    db.commit()
    db.refresh(card)
    return card


@router.get("", response_model=list[CardOut])
def list_cards(
    q: str = "",
    status: str = "",
    sport: str = "",
    db: Session = Depends(get_db),
):
    """List inventory with optional text search and status/sport filters."""
    query = db.query(Card)
    if status:
        query = query.filter(Card.status == status)
    if sport:
        query = query.filter(Card.sport == sport)
    if q:
        like = f"%{q.lower()}%"
        query = query.filter(
            func.lower(Card.player).like(like)
            | func.lower(Card.set_name).like(like)
            | func.lower(Card.brand).like(like)
            | func.lower(Card.team).like(like)
            | func.lower(Card.card_id).like(like)
            | func.lower(Card.year).like(like)
        )
    return query.order_by(Card.id.desc()).all()


_EXPORT_HEADERS = [
    "Card ID", "Player", "Year", "Brand", "Set", "Card #", "Variation",
    "Team", "Sport", "Rookie", "Condition", "Status",
    "Purchase Date", "Purchase Source", "Purchase Price",
    "Sale Date", "Sale Platform", "Sale Price", "Sale Fees", "Sale Shipping",
    "Net Profit", "Est. Value", "Notes", "Front Image", "Back Image",
]


def _card_row(c: Card) -> list:
    def num(v):
        return v if v is not None else ""
    return [
        c.card_id, c.player, c.year, c.brand, c.set_name, c.card_number,
        c.variation, c.team, c.sport, c.is_rookie, c.condition, c.status,
        c.purchase_date, c.purchase_source, num(c.purchase_price),
        c.sale_date, c.sale_platform, num(c.sale_price), num(c.sale_fees),
        num(c.sale_shipping), num(c.net_profit), num(c.estimated_value),
        c.notes, c.front_image, c.back_image,
    ]


@router.get("/export.csv")
def export_csv(db: Session = Depends(get_db)):
    """Export the whole inventory as CSV (Google Sheets friendly)."""
    cards = db.query(Card).order_by(Card.card_id).all()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_EXPORT_HEADERS)
    for c in cards:
        writer.writerow(_card_row(c))
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=card_inventory.csv"},
    )


@router.get("/export.xlsx")
def export_xlsx(db: Session = Depends(get_db)):
    """Export the whole inventory as an Excel .xlsx workbook (also serves as the import template)."""
    from openpyxl import Workbook

    cards = db.query(Card).order_by(Card.card_id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(_EXPORT_HEADERS)
    for c in cards:
        ws.append(_card_row(c))
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=card_inventory.xlsx"},
    )


@router.get("/{card_pk}", response_model=CardOut)
def get_card(card_pk: int, db: Session = Depends(get_db)):
    card = db.get(Card, card_pk)
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    return card


@router.put("/{card_pk}", response_model=CardOut)
def update_card(card_pk: int, payload: CardUpdate, db: Session = Depends(get_db)):
    card = db.get(Card, card_pk)
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    data = payload.model_dump(exclude_unset=True)
    if "status" in data and data["status"] not in VALID_STATUSES:
        data.pop("status")
    for key, value in data.items():
        setattr(card, key, value)
    db.commit()
    db.refresh(card)
    return card


@router.delete("/{card_pk}")
def delete_card(card_pk: int, db: Session = Depends(get_db)):
    card = db.get(Card, card_pk)
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    db.delete(card)
    db.commit()
    return {"ok": True}
