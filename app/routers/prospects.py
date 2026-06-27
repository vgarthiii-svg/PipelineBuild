import csv
import io
import re
from urllib.parse import urlparse
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional

from app.database import get_db
from app.models import Prospect, Client, PipelineEntry, Relationship
from app.schemas import ProspectCreate, ProspectOut, BulkProspectCreate

router = APIRouter(prefix="/api/prospects", tags=["prospects"])


def _add_prospect_to_all_pipelines(prospect, db, source="Auto-added"):
    """
    Ensure this prospect appears in every client's pipeline (active or not).
    Used by every prospect-creation path so Layer 1 and the pipeline never drift.
    Skips clients where this prospect would be the client itself (by name match).
    Returns count added.
    """
    clients = db.query(Client).all()
    added = 0
    for c in clients:
        if c.name.lower() == prospect.name.lower():
            continue
        existing = db.query(PipelineEntry).filter(
            PipelineEntry.client_id == c.id,
            PipelineEntry.prospect_id == prospect.id,
        ).first()
        if existing:
            continue
        db.add(PipelineEntry(
            client_id=c.id,
            prospect_id=prospect.id,
            source=source,
        ))
        added += 1
    return added


@router.delete("/{prospect_id}/cascade-everywhere")
def delete_prospect_everywhere(prospect_id: int, db: Session = Depends(get_db)):
    """
    Nuclear delete: remove this company from the whole system.
    Wipes: prospect record, all pipeline entries across all clients,
    matching client record (if any), business profile, behavior profile,
    relationships, intent signals.
    """
    from app.models import (CriterionScore, IntroPackage, ActivityLog, ScoringCriterion,
                            IdentityFilter, BusinessProfile, BehaviorProfile,
                            Relationship, IntentSignal)

    prospect = db.query(Prospect).filter(Prospect.id == prospect_id).first()
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospect not found")

    name = prospect.name

    # Wipe all pipeline entries referencing this prospect (across every client)
    entry_ids = [e.id for e in db.query(PipelineEntry.id).filter(PipelineEntry.prospect_id == prospect.id).all()]
    if entry_ids:
        db.query(CriterionScore).filter(CriterionScore.pipeline_entry_id.in_(entry_ids)).delete(synchronize_session=False)
        db.query(IntroPackage).filter(IntroPackage.pipeline_entry_id.in_(entry_ids)).delete(synchronize_session=False)
        db.query(ActivityLog).filter(ActivityLog.pipeline_entry_id.in_(entry_ids)).delete(synchronize_session=False)
        db.query(PipelineEntry).filter(PipelineEntry.prospect_id == prospect.id).delete(synchronize_session=False)

    # Wipe profiles and relationships
    db.query(BusinessProfile).filter(BusinessProfile.prospect_id == prospect.id).delete(synchronize_session=False)
    db.query(BehaviorProfile).filter(BehaviorProfile.prospect_id == prospect.id).delete(synchronize_session=False)
    db.query(Relationship).filter(Relationship.prospect_id == prospect.id).delete(synchronize_session=False)
    db.query(IntentSignal).filter(IntentSignal.prospect_id == prospect.id).delete(synchronize_session=False)

    # Delete matching client record if one exists
    client = db.query(Client).filter(Client.name.ilike(name)).first()
    if client:
        client_entry_ids = [e.id for e in db.query(PipelineEntry.id).filter(PipelineEntry.client_id == client.id).all()]
        if client_entry_ids:
            db.query(CriterionScore).filter(CriterionScore.pipeline_entry_id.in_(client_entry_ids)).delete(synchronize_session=False)
            db.query(IntroPackage).filter(IntroPackage.pipeline_entry_id.in_(client_entry_ids)).delete(synchronize_session=False)
            db.query(ActivityLog).filter(ActivityLog.pipeline_entry_id.in_(client_entry_ids)).delete(synchronize_session=False)
            db.query(PipelineEntry).filter(PipelineEntry.client_id == client.id).delete(synchronize_session=False)
        db.query(ScoringCriterion).filter(ScoringCriterion.client_id == client.id).delete(synchronize_session=False)
        db.query(IdentityFilter).filter(IdentityFilter.client_id == client.id).delete(synchronize_session=False)
        db.delete(client)

    db.delete(prospect)
    db.commit()
    return {"status": "deleted_everywhere", "name": name, "removed_client": bool(client)}


@router.get("/", response_model=List[ProspectOut])
def list_prospects(
    search: Optional[str] = None,
    type: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Prospect)
    if search:
        q = q.filter(Prospect.name.ilike(f"%{search}%"))
    if type:
        q = q.filter(Prospect.type == type)
    return q.order_by(Prospect.name).all()


@router.get("/available-as-clients")
def list_prospects_not_clients(db: Session = Depends(get_db)):
    """Return all prospects that are NOT already clients (by name match)."""
    client_names = [c.name.lower() for c in db.query(Client.name).all()]
    prospects = db.query(Prospect).order_by(Prospect.name).all()
    return [
        {"id": p.id, "name": p.name, "type": p.type}
        for p in prospects
        if p.name.lower() not in client_names
    ]


@router.post("/", response_model=ProspectOut)
def create_prospect(data: ProspectCreate, db: Session = Depends(get_db)):
    from app.services import profile_generator
    prospect = Prospect(**data.dict())
    db.add(prospect)
    db.flush()
    try:
        profile_generator.upsert_profile(prospect, db, run_enrichment=True)
    except Exception as e:
        print(f"[PROSPECTS] Profile generation failed: {e}")
    # Sync: every new prospect must appear in every client's pipeline
    _add_prospect_to_all_pipelines(prospect, db, source="Created via POST /api/prospects")
    db.commit()
    db.refresh(prospect)
    return prospect


@router.post("/bulk", response_model=List[ProspectOut])
def bulk_create_prospects(data: BulkProspectCreate, db: Session = Depends(get_db)):
    """Create prospects from a list of company names. Auto-generates profiles + types. Skips duplicates."""
    from app.services import profile_generator, type_inference
    created = []
    for name in data.names:
        name = name.strip()
        if not name:
            continue
        existing = db.query(Prospect).filter(Prospect.name.ilike(name)).first()
        if existing:
            created.append(existing)
            continue
        # Infer accurate type from HubSpot + Apollo + keyword matching (all free)
        inferred_type = type_inference.infer_type(name=name)
        prospect = Prospect(name=name, type=inferred_type)
        db.add(prospect)
        db.flush()
        # Auto-generate profile (free Layer 1 + Layer 2 if keys set)
        try:
            profile_generator.upsert_profile(prospect, db, run_enrichment=True)
        except Exception as e:
            print(f"[PROSPECTS] Profile generation failed for {name}: {e}")
        # Sync to every client's pipeline
        _add_prospect_to_all_pipelines(prospect, db, source="Bulk prospect create")
        created.append(prospect)
    db.commit()
    for p in created:
        db.refresh(p)
    return created


@router.post("/import-csv")
async def import_csv(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import prospects from CSV. Expected columns:
    Partner Name, Partner Type, Region, Tier (Guidewire format)
    OR: name, type, domain, hq_city, hq_state
    """
    content = await file.read()
    text = content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    for row in reader:
        # Support both Guidewire format and generic format
        name = row.get("Partner Name") or row.get("name") or ""
        name = name.strip()
        if not name:
            continue

        existing = db.query(Prospect).filter(Prospect.name.ilike(name)).first()
        if existing:
            skipped += 1
            continue

        prospect = Prospect(
            name=name,
            type=row.get("Partner Type") or row.get("type"),
            domain=row.get("domain"),
            hq_city=row.get("hq_city"),
            hq_state=row.get("hq_state"),
        )
        db.add(prospect)
        db.flush()
        # Sync to every client's pipeline
        _add_prospect_to_all_pipelines(prospect, db, source="CSV import")
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped}


def _domain_from_url(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    url = url.strip()
    if not url:
        return None
    if not re.match(r"^https?://", url, re.I):
        url = "http://" + url
    try:
        host = urlparse(url).hostname or ""
        return host.lower().lstrip("www.") or None
    except Exception:
        return None


@router.post("/import-xlsx")
async def import_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import prospects + relationships from a HubSpot-style xlsx (Pipeline Book).
    Expected columns (header row 1): First Name, Last Name, Full Name, Email,
    Company, Title, LinkedIn URL, Company Website, Phone, Partner Type,
    Pipeline Stage, BD Team Lead, Priority, Lifecycle Stage, My Contact,
    HubSpot ID, Source Notes.

    One Prospect per unique Company (created if missing, never overwritten),
    one Relationship per contact row (deduped on email within a prospect).
    Default relationship score = 2. Source = "pipeline_book". This is a pure
    ingest: no profile generation is triggered.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read xlsx: {e}")

    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
    except StopIteration:
        return {"prospects_created": 0, "relationships_created": 0, "rows_skipped": 0}

    idx = {name: i for i, name in enumerate(header)}

    def cell(row, col_name):
        i = idx.get(col_name)
        if i is None or i >= len(row):
            return None
        v = row[i]
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    prospects_created = 0
    relationships_created = 0
    rows_skipped = 0

    for row in rows:
        if not any(c is not None and str(c).strip() for c in row):
            continue

        company = cell(row, "Company")
        if not company:
            rows_skipped += 1
            continue

        prospect = db.query(Prospect).filter(Prospect.name.ilike(company)).first()
        if not prospect:
            prospect = Prospect(
                name=company,
                type=cell(row, "Partner Type"),
                website=cell(row, "Company Website"),
                domain=_domain_from_url(cell(row, "Company Website")),
            )
            db.add(prospect)
            db.flush()
            prospects_created += 1

        email = cell(row, "Email")
        full_name = cell(row, "Full Name") or " ".join(
            filter(None, [cell(row, "First Name"), cell(row, "Last Name")])
        ) or None

        if email:
            existing = (
                db.query(Relationship)
                .filter(
                    Relationship.prospect_id == prospect.id,
                    Relationship.contact_email.ilike(email),
                )
                .first()
            )
            if existing:
                rows_skipped += 1
                continue
        elif full_name:
            existing = (
                db.query(Relationship)
                .filter(
                    Relationship.prospect_id == prospect.id,
                    Relationship.contact_name.ilike(full_name),
                )
                .first()
            )
            if existing:
                rows_skipped += 1
                continue
        else:
            rows_skipped += 1
            continue

        context_parts = []
        for label, col in [
            ("Pipeline Stage", "Pipeline Stage"),
            ("Lifecycle", "Lifecycle Stage"),
            ("Priority", "Priority"),
            ("My Contact", "My Contact"),
            ("BD Lead", "BD Team Lead"),
            ("HubSpot ID", "HubSpot ID"),
            ("Notes", "Source Notes"),
        ]:
            val = cell(row, col)
            if val:
                context_parts.append(f"{label}: {val}")

        rel = Relationship(
            prospect_id=prospect.id,
            contact_name=full_name,
            contact_title=cell(row, "Title"),
            contact_email=email,
            contact_linkedin=cell(row, "LinkedIn URL"),
            score=2,
            context=" | ".join(context_parts) or None,
            source="pipeline_book",
        )
        db.add(rel)
        relationships_created += 1

    db.commit()
    return {
        "prospects_created": prospects_created,
        "relationships_created": relationships_created,
        "rows_skipped": rows_skipped,
    }


@router.get("/{prospect_id}", response_model=ProspectOut)
def get_prospect(prospect_id: int, db: Session = Depends(get_db)):
    prospect = db.query(Prospect).filter(Prospect.id == prospect_id).first()
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospect not found")
    return prospect


@router.put("/{prospect_id}/type")
def update_prospect_type(prospect_id: int, new_type: str, db: Session = Depends(get_db)):
    """
    Update a prospect's type and automatically re-infer their behavior profile.
    Also rescores any pipeline entries that reference this prospect.
    """
    from app.models import BehaviorProfile, PipelineEntry
    from app.routers.behavior import TYPE_BEHAVIOR_MAP
    from datetime import datetime

    prospect = db.query(Prospect).filter(Prospect.id == prospect_id).first()
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospect not found")

    old_type = prospect.type
    prospect.type = new_type
    db.flush()

    # Re-infer behavior profile based on new type
    if new_type in TYPE_BEHAVIOR_MAP:
        bp = db.query(BehaviorProfile).filter(BehaviorProfile.prospect_id == prospect_id).first()
        defaults = TYPE_BEHAVIOR_MAP[new_type]
        if bp:
            # Only overwrite if it was auto-inferred, never overwrite manually-set profiles
            if bp.assessed_by in ("heuristic_auto", "heuristic_bulk_infer", "seed"):
                for k, v in defaults.items():
                    setattr(bp, k, v)
                bp.assessed_by = "heuristic_auto"
                bp.assessed_at = datetime.utcnow()
                bp.partnership_history = f"Auto-inferred from type: {new_type}. Edit via Behavior tab."
        else:
            bp = BehaviorProfile(
                prospect_id=prospect_id,
                assessed_by="heuristic_auto",
                assessed_at=datetime.utcnow(),
                partnership_history=f"Auto-inferred from type: {new_type}. Edit via Behavior tab.",
                **defaults,
            )
            db.add(bp)

    db.commit()

    # Rescore all client pipelines that contain this prospect
    from app.routers.clients import _auto_score_pipeline
    entries = db.query(PipelineEntry).filter(PipelineEntry.prospect_id == prospect_id).all()
    client_ids = list({e.client_id for e in entries})
    for cid in client_ids:
        _auto_score_pipeline(cid, db)
    db.commit()

    return {
        "prospect_id": prospect_id,
        "old_type": old_type,
        "new_type": new_type,
        "behavior_profile_updated": new_type in TYPE_BEHAVIOR_MAP,
        "pipelines_rescored": len(client_ids),
    }


@router.put("/{prospect_id}", response_model=ProspectOut)
def update_prospect(prospect_id: int, data: ProspectCreate, db: Session = Depends(get_db)):
    prospect = db.query(Prospect).filter(Prospect.id == prospect_id).first()
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospect not found")
    for key, val in data.dict(exclude_unset=True).items():
        setattr(prospect, key, val)
    db.commit()
    db.refresh(prospect)
    return prospect
